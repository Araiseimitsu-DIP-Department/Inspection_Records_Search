"""検索結果の Excel 出力（openpyxl）。"""

from __future__ import annotations

import datetime as dt
from decimal import Decimal
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font


def _cell_value(v: Any) -> Any:
    if v is None:
        return None
    if isinstance(v, dt.datetime):
        return v.replace(tzinfo=None) if v.tzinfo else v
    if isinstance(v, dt.date):
        return v
    if isinstance(v, Decimal):
        return float(v)
    if isinstance(v, bytes):
        return v.decode(errors="replace")
    return v


def export_to_xlsx(
    db_path: str | None,
    filename: str,
    headers: list[str],
    rows: list[tuple],
    output_dir: Path | None = None,
    output_path: Path | None = None,
) -> Path:
    """
    output_path があればそのパスへ保存する。
    無ければ output_dir、db_path と同じディレクトリ、カレントディレクトリの順で保存する。
    戻り値は保存したパス。
    """
    if output_path is not None:
        out_path = output_path
        out_path.parent.mkdir(parents=True, exist_ok=True)
    elif output_dir is not None:
        out_dir = output_dir
        out_dir.mkdir(parents=True, exist_ok=True)
        out_path = out_dir / filename
    elif db_path:
        out_dir = Path(db_path).resolve().parent
        out_dir.mkdir(parents=True, exist_ok=True)
        out_path = out_dir / filename
    else:
        out_dir = Path.cwd()
        out_dir.mkdir(parents=True, exist_ok=True)
        out_path = out_dir / filename

    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    bold = Font(bold=True)
    for col, name in enumerate(headers, start=1):
        c = ws.cell(row=1, column=col, value=name)
        c.font = bold

    for r_idx, row in enumerate(rows, start=2):
        for c_idx, val in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=_cell_value(val))

    wb.save(out_path)
    return out_path
